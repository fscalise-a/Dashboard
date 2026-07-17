"""
ECO GO Dashboard - Refresh de datos
====================================
Re-lee los Excel originales y regenera los archivos JSON/JS del dashboard.

Uso:
  - Doble click a refresh.bat (recomendado)
  - O desde terminal: python refresh.py

Los archivos Excel NO se modifican (read-only).
"""

import os
import sys
import json
import re
import traceback
from datetime import datetime

# Asegurar consola con soporte unicode en Windows
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

# =====================================================================
#  CONFIGURACIÓN — Editá estas rutas si los Excel cambian de ubicación
# =====================================================================
BASE_EXCEL = "/sessions/sweet-dreamy-wright/mnt/BD"

EXCEL_PATHS = {
    "ipc":         os.path.join(BASE_EXCEL, "Precios", "IPC TODESCA.xlsx"),
    "gd":          os.path.join(BASE_EXCEL, "Precios", "Gráficos de dispersión - copia - copia.xlsx"),
    "cm":          os.path.join(BASE_EXCEL, "Precios", "CM - DB.xlsx"),  # opcional
    "empleo":      os.path.join(BASE_EXCEL, "Empleo",  "Empleo_nuevo.xlsx"),
    "salarios":    os.path.join(BASE_EXCEL, "Empleo",  "Salarios.xlsx"),
    "tcr_bandas":  os.path.join(BASE_EXCEL, "Tipo de Cambio", "TCR bandas.xlsx"),
    "rofex":       os.path.join(BASE_EXCEL, "Tipo de Cambio", "Rofex.xlsx"),
    "com3500":     os.path.join(BASE_EXCEL, "Tipo de Cambio", "com3500.xls"),  # TCN A3500 historico
}

# Monitor mundial — no es Excel, es un .js con datos del monitor externo
MONITOR_MUNDIAL_JS = os.path.join(BASE_EXCEL, "Internacional", "Monitor mundial", "data", "monitor-data.js")

# Directorio del dashboard = donde está este script
DASHBOARD_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(DASHBOARD_DIR, "assets", "data")

# =====================================================================
#  Helpers
# =====================================================================

class Status:
    def __init__(self):
        self.results = []
    def ok(self, name, detail=""):
        self.results.append(("OK", name, detail))
        print(f"  [OK]   {name}" + (f" — {detail}" if detail else ""))
    def warn(self, name, detail=""):
        self.results.append(("WARN", name, detail))
        print(f"  [---]  {name}" + (f" — {detail}" if detail else ""))
    def fail(self, name, detail=""):
        self.results.append(("FAIL", name, detail))
        print(f"  [FAIL] {name}" + (f" — {detail}" if detail else ""))

def save_data(name, data):
    """Guarda como JSON + JS con la variable global esperada."""
    json_path = os.path.join(DATA_DIR, f"{name}.json")
    js_path   = os.path.join(DATA_DIR, f"{name}.js")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, default=str)
    var_name = {"precios":"PRECIOS_DATA","empleo":"EMPLEO_DATA","salarios":"SALARIOS_DATA","tipo-cambio":"TC_DATA"}.get(name, name.upper()+"_DATA")
    js = f"// Datos de {name} - regenerado por refresh.py el {datetime.now().strftime('%Y-%m-%d %H:%M')}\nwindow.{var_name} = {json.dumps(data, ensure_ascii=False, default=str)};\n"
    with open(js_path, "w", encoding="utf-8") as f:
        f.write(js)
    return os.path.getsize(json_path)

def load_existing(name):
    json_path = os.path.join(DATA_DIR, f"{name}.json")
    if os.path.exists(json_path):
        try:
            with open(json_path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None
    return None

def col_idx(letter):
    n = 0
    for c in letter:
        n = n * 26 + ord(c) - 64
    return n

def fmt_n(v):
    return v if isinstance(v, (int, float)) else None

# =====================================================================
#  PRECIOS
# =====================================================================
def extract_precios(status):
    import openpyxl

    if not os.path.exists(EXCEL_PATHS["ipc"]):
        status.fail("Precios", f"No se encontró: {EXCEL_PATHS['ipc']}")
        return None
    if not os.path.exists(EXCEL_PATHS["gd"]):
        status.fail("Precios - GD", f"No se encontró: {EXCEL_PATHS['gd']}")
        return None

    data = {}

    # ---- IPC último + series ----
    wb = openpyxl.load_workbook(EXCEL_PATHS["ipc"], data_only=True)
    ws = wb["1. Nuevo IPC Nacional"]

    last = None
    for r in range(ws.max_row, 0, -1):
        if ws.cell(r, 20).value is not None:  # col T (nucleo m/m)
            last = r
            break

    def g(r, col):
        return ws.cell(r, col_idx(col)).value

    # Último mes
    fecha = g(last, 'A')
    data['ultimo'] = {
        "fecha": fecha.strftime("%Y-%m-%d") if isinstance(fecha, datetime) else str(fecha),
        "general_mm":   fmt_n(g(last, 'AF')),
        "general_ia":   fmt_n(g(last, 'AT')),
        "nucleo_mm":    fmt_n(g(last, 'T')),
        "nucleo_ia":    fmt_n(g(last, 'U')),
        "estacional_mm": fmt_n(g(last, 'P')),
        "estacional_ia": fmt_n(g(last, 'BV')),
        "regulados_mm": fmt_n(g(last, 'V')),
        "regulados_ia": fmt_n(g(last, 'W'))
    }

    # Series 24 meses
    series = []
    for r in range(max(6, last - 24), last + 1):
        fecha = g(r, 'A')
        if not isinstance(fecha, datetime): continue
        series.append({
            "fecha": fecha.strftime("%Y-%m-%d"),
            "general_mm": fmt_n(g(r, 'AF')), "general_ia": fmt_n(g(r, 'AT')),
            "nucleo_mm": fmt_n(g(r, 'T')),   "nucleo_ia": fmt_n(g(r, 'U')),
            "estacional_mm": fmt_n(g(r, 'P')), "estacional_ia": fmt_n(g(r, 'BV')),
            "regulados_mm": fmt_n(g(r, 'V')), "regulados_ia": fmt_n(g(r, 'W'))
        })
    data['series_24m'] = series

    # ---- Chart 21: '1. Nuevo IPC Nacional'!A17:A114, DJ, DK, AT ----
    def range_get(sheet, col, r1, r2):
        out = []
        for r in range(r1, r2+1):
            v = sheet.cell(r, col_idx(col)).value
            out.append(v)
        return out

    cats = range_get(ws, 'A', 17, 114)
    chart21 = []
    for i, c in enumerate(cats):
        if c is None: continue
        chart21.append({
            "fecha": c.strftime("%Y-%m-%d") if isinstance(c, datetime) else str(c),
            "mensual": range_get(ws,'DJ',17,114)[i],
            "promedio_anual": range_get(ws,'DK',17,114)[i],
            "var_ia": range_get(ws,'AT',17,114)[i]
        })
    data['chart21'] = chart21

    # ---- Chart 23: A6:A115, AF, AT, U, W, DK ----
    cats5 = range_get(ws, 'A', 6, 115)
    chart23 = []
    var_men = range_get(ws,'AF',6,115)
    var_ia5 = range_get(ws,'AT',6,115)
    nucleo_ia = range_get(ws,'U',6,115)
    w_series = range_get(ws,'W',6,115)
    prom_men = range_get(ws,'DK',6,115)
    for i, c in enumerate(cats5):
        if c is None: continue
        chart23.append({
            "fecha": c.strftime("%Y-%m-%d") if isinstance(c, datetime) else str(c),
            "var_men": var_men[i], "var_ia": var_ia5[i],
            "nucleo_ia": nucleo_ia[i], "regulados_ia": w_series[i],
            "prom_men": prom_men[i]
        })
    data['chart23'] = chart23

    # ---- Chart 22 (RPM): '4. Proyecciones'!B100:B201, C/E/F/H ----
    ws2 = wb["4. Proyecciones"]
    chart22 = []
    for r in range(100, 202):
        fecha = ws2.cell(r, 2).value
        if fecha is None: continue
        chart22.append({
            "fecha": fecha.strftime("%Y-%m-%d") if isinstance(fecha, datetime) else str(fecha),
            "rpm_mm": fmt_n(ws2.cell(r, 3).value),
            "ipc_gba": fmt_n(ws2.cell(r, 5).value),
            "ipc_nac": fmt_n(ws2.cell(r, 6).value),
            "rpm_ia": fmt_n(ws2.cell(r, 8).value)
        })
    data['chart22'] = chart22

    # ---- Chart 5: Gráficos de dispersión - Hoja3 ----
    wb_gd = openpyxl.load_workbook(EXCEL_PATHS["gd"], data_only=True)
    ws_gd = wb_gd["Hoja3"]
    CATS = {
        'Pan y cereales':'Bienes','Carnes y derivados':'Bienes',
        'Leche, productos lácteos y huevos':'Bienes','Aceites, grasas y manteca':'Bienes',
        'Frutas y verduras':'Bienes','Azúcar, dulces, chocolates':'Bienes',
        'Bebidas':'Bienes','Tabaco':'Bienes','Indumentaria':'Bienes',
        'Medicamentos':'Bienes','Medicamentos ':'Bienes',
        'Adquisición de vehículos':'Bienes','Combustibles':'Bienes',
        'Electrodomésticos':'Bienes','Bienes y servicios para la conservación del hogar':'Bienes',
        'Bienes':'Bienes',
        'Vivienda':'Servicios','Alquiler de la vivienda y gastos conexos':'Servicios',
        'Prepagas':'Servicios','Servicios recreativos':'Servicios','Servicios recreativos ':'Servicios',
        'Educación':'Servicios','Restaurantes':'Servicios','Servicios':'Servicios',
        'Electricidad, gas y otros combustibles':'Regulados','Transporte público':'Regulados',
        'Servicios  de telefonía e internet':'Regulados','Servicios de telefonía e internet':'Regulados',
        'Tarifa de Agua':'Regulados','Regulados':'Regulados',
        'Salarios Formales Privados':'Salarios','Salarios Informales':'Salarios',
        'Salarios Informales ':'Salarios','Jubilación Mínima c/Bono':'Salarios',
        'Jubilación s/Bono':'Salarios','Salarios Formales Públicos':'Salarios',
        'AUH':'Salarios','Salario de Cuentapropistas':'Salarios',
        'Dólar CCL':'Dólar','Dólar Oficial':'Dólar',
        'ICC(Materiales p/construcción)':'Construcción','ICC (Nivel Gral.)':'Construcción',
        'ICC (Mano de Obra)':'Construcción','ICC (Gastos Generales)':'Construcción',
        'IPIM (Nivel General)':'Mayorista','IPIM (Nacionales)':'Mayorista',
        'IPIM (Importados)':'Mayorista'
    }
    points = []
    for r in list(range(3, 42)) + list(range(44, 48)):
        label = ws_gd.cell(r, 1).value
        x = ws_gd.cell(r, 6).value
        y = ws_gd.cell(r, 7).value
        if label and x is not None and y is not None:
            cat = CATS.get(label, 'Otros') if not isinstance(label, str) else CATS.get(label.strip(), 'Otros')
            points.append({"label": str(label).strip(), "categoria": cat, "x": x, "y": y})
    data['chart5'] = {
        "title": "Precios relativos: cuánto corregimos y cuánto falta corregir",
        "x_label": str(ws_gd.cell(1, 6).value or ""),
        "y_label": str(ws_gd.cell(1, 7).value or ""),
        "points": points
    }

    # ---- Proyección RPM (CM - DB hoja 2605) - opcional ----
    if os.path.exists(EXCEL_PATHS["cm"]):
        try:
            wb_cm = openpyxl.load_workbook(EXCEL_PATHS["cm"], data_only=True, read_only=True)
            # Buscar la hoja más reciente con formato YYMM
            yymm_sheets = sorted([s for s in wb_cm.sheetnames if s.isdigit() and len(s) == 4], reverse=True)
            if yymm_sheets:
                hoja = yymm_sheets[0]
                ws_cm = wb_cm[hoja]
                rows = list(ws_cm.iter_rows(min_row=80, max_row=100, max_col=6, values_only=True))
                proy = {
                    "titulo": rows[2][0] if len(rows) > 2 else "",
                    "subtitulo": rows[3][0] if len(rows) > 3 else "",
                    "header_periodo": f"Proyección — RPM Eco Go (hoja {hoja})",
                    "filas": []
                }
                for ri in range(7, 19):
                    if ri >= len(rows): break
                    r = rows[ri]
                    if r[0] and r[1] is not None:
                        proy["filas"].append({
                            "capitulo": str(r[0]),
                            "mensual": r[1],
                            "anual": r[2],
                            "acumulada": r[3]
                        })
                data['proyeccion'] = proy
                status.ok("Precios - Proyección RPM", f"hoja {hoja}, {len(proy['filas'])} filas")
            else:
                _keep_old_proy(data, status)
        except Exception as e:
            status.warn("Precios - Proyección RPM", f"error: {e}; mantengo datos previos")
            _keep_old_proy(data, status)
    else:
        _keep_old_proy(data, status)

    return data

def _keep_old_proy(data, status):
    """Si no hay CM-DB, conservar la proyección anterior del JSON existente."""
    old = load_existing("precios")
    if old and "proyeccion" in old:
        data["proyeccion"] = old["proyeccion"]
        status.warn("Precios - Proyección RPM", "CM-DB no encontrado, mantengo dato previo")
    else:
        status.warn("Precios - Proyección RPM", "sin CM-DB ni dato previo")

# =====================================================================
#  EMPLEO (EPH + Cuadro Trim + SIPA + Provincias)
# =====================================================================
def extract_empleo(status):
    import openpyxl
    if not os.path.exists(EXCEL_PATHS["empleo"]):
        status.fail("Empleo", f"No se encontró: {EXCEL_PATHS['empleo']}")
        return None

    wb = openpyxl.load_workbook(EXCEL_PATHS["empleo"], data_only=True)
    data = {}

    # ---- Tasas EPH (sección principal R8-R124, cols B,E,H,K,N,Q,T) ----
    ws = wb["Tasas EPH"]
    eph = []
    for r in range(8, 125):
        fecha = ws.cell(r, 1).value
        if fecha is None: continue
        if isinstance(fecha, datetime):
            fecha_str = fecha.strftime("%Y-%m-%d")
            ord_idx = (fecha.year * 4) + (fecha.month - 1) // 3
            label = fecha.strftime("%b-%Y")
        elif isinstance(fecha, str):
            m = re.match(r'(IV|III|II|I)-(\d{2,4})', fecha.strip())
            if m:
                qmap = {'I':1,'II':2,'III':3,'IV':4}
                q = qmap[m.group(1)]
                yy = int(m.group(2))
                if yy < 100: yy += 2000
                label = fecha.strip()
            else:
                m2 = re.match(r'(\d)\s*º?\s*trim\s*(\d{4})', fecha.strip(), re.IGNORECASE)
                if not m2: continue
                q = int(m2.group(1))
                yy = int(m2.group(2))
                label = f"{['','I','II','III','IV'][q]}-{str(yy)[2:]}"
            fecha_str = f"{yy}-{(q-1)*3+1:02d}-01"
            ord_idx = (yy * 4) + (q - 1)
        else:
            continue
        eph.append({
            "fecha": fecha_str, "ord": ord_idx, "label": label,
            "actividad":     fmt_n(ws.cell(r, 2).value),
            "empleo":        fmt_n(ws.cell(r, 5).value),
            "desocup":       fmt_n(ws.cell(r, 8).value),
            "ocup_dem":      fmt_n(ws.cell(r, 11).value),
            "subocup":       fmt_n(ws.cell(r, 14).value),
            "subocup_dem":   fmt_n(ws.cell(r, 17).value),
            "subocup_nodem": fmt_n(ws.cell(r, 20).value)
        })
    eph.sort(key=lambda x: x['ord'])
    data['eph'] = eph
    data['eph_ultimo'] = eph[-1] if eph else None

    # ---- Cuadro empleo trim ----
    ws = wb["Cuadro empleo trim"]
    COLS_TRIM = [(3,'2012 (est.)'), (4,'II-18'), (5,'II-23'), (7,'IV-23'), (8,'I-24'),
                 (9,'II-24'), (10,'IV-24'), (12,'I-25'), (13,'III-25'), (14,'IV-25')]
    trim = {"periodos": [c[1] for c in COLS_TRIM], "filas": []}
    for r in range(5, 20):
        cat = ws.cell(r, 2).value
        if not cat: continue
        valores = []
        for col_idx_n, _ in COLS_TRIM:
            v = ws.cell(r, col_idx_n).value
            valores.append(v if isinstance(v, (int, float)) else None)
        trim["filas"].append({"categoria": str(cat).strip(), "valores": valores})
    data['trim'] = trim

    # ---- Cuadro SIPA ext (2) ----
    ws = wb["Cuadro SIPA ext (2)"]
    COLS_SIPA = [
        (3,'Jun-12','en miles','stock'),(4,'Ago-18*','en miles','stock'),
        (5,'Dic-23*','en miles','stock'),(6,'Feb-25*','en miles','stock'),
        (9,'Ene-26*','en miles','stock'),(12,'Feb-26*','en miles','stock'),
        (13,'Dif. (6)-(1)','vs Jun-12','diff'),(14,'Dif. (6)-(2)','vs Ago-18','diff'),
        (15,'Dif. (6)-(3)','vs Dic-23','diff'),(16,'Dif. (6)-(4)','vs Feb-25','diff'),
        (17,'Dif. (6)-(5)','vs Ene-26','diff')
    ]
    sipa = {"cols": [{"header":c[1],"sub":c[2],"tipo":c[3]} for c in COLS_SIPA], "filas":[]}
    for r in range(4, 13):
        cat = ws.cell(r, 2).value
        if not cat: continue
        valores = []
        for col_idx_n, _, _, _ in COLS_SIPA:
            v = ws.cell(r, col_idx_n).value
            valores.append(v if isinstance(v, (int, float)) else None)
        sipa["filas"].append({"categoria": str(cat).strip(), "valores": valores})
    data['sipa'] = sipa

    # ---- Provincias (Hoja7) ----
    ws = wb["Hoja7"]
    provincias = {"fechas": [], "datos": []}
    for c in range(3, 9):
        v = ws.cell(3, c).value
        if isinstance(v, datetime):
            provincias["fechas"].append(v.strftime("%Y-%m"))
    for r in range(5, 30):
        prov = ws.cell(r, 2).value
        if not prov: continue
        vals = []
        for c in range(3, 9):
            v = ws.cell(r, c).value
            vals.append(v if isinstance(v, (int, float)) else None)
        if any(v is not None for v in vals):
            provincias["datos"].append({"provincia": str(prov).strip(), "valores": vals})
    data['provincias'] = provincias

    return data

# =====================================================================
#  SALARIOS
# =====================================================================
def extract_salarios(status):
    import openpyxl
    if not os.path.exists(EXCEL_PATHS["salarios"]):
        status.fail("Salarios", f"No se encontró: {EXCEL_PATHS['salarios']}")
        return None

    wb = openpyxl.load_workbook(EXCEL_PATHS["salarios"], data_only=True)
    data = {}

    # ---- G Sal real: 1.1 INDEC A186:A298, CE/CF/CH/CI ----
    ws = wb["1.1 INDEC"]
    sal_real = []
    for r in range(186, 299):
        fecha = ws.cell(r, 1).value
        if not isinstance(fecha, datetime): continue
        priv = fmt_n(ws.cell(r, 83).value)  # CE
        pub  = fmt_n(ws.cell(r, 84).value)  # CF
        nor  = fmt_n(ws.cell(r, 86).value)  # CH
        tot  = fmt_n(ws.cell(r, 87).value)  # CI
        if any(v is not None for v in [priv, pub, nor, tot]):
            sal_real.append({
                "fecha": fecha.strftime("%Y-%m-%d"),
                "label": fecha.strftime("%b-%y"),
                "priv": priv, "pub": pub, "nor": nor, "tot": tot
            })
    data['sal_real'] = sal_real

    # ---- Cuadro INDEC 1.3 ----
    ws = wb["1.3 Cuadro INDEC"]
    fecha_actual = ws.cell(3, 2).value
    fecha_compar = ws.cell(3, 11).value
    def iso(v):
        if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
        return str(v) if v else ''
    cuadro = {
        "fecha_actual": iso(fecha_actual),
        "fecha_comparacion": iso(fecha_compar),
        "headers": [],
        "filas": []
    }
    for c in range(3, 11):
        v = ws.cell(4, c).value
        cuadro["headers"].append(str(v).replace('\n',' ').strip() if v else '')
    for r in range(5, 10):
        cat = ws.cell(r, 2).value
        if not cat: continue
        valores = []
        for c in range(3, 11):
            v = ws.cell(r, c).value
            valores.append(v if isinstance(v, (int, float)) else None)
        cuadro["filas"].append({"categoria": str(cat).strip(), "valores": valores})
    data['cuadro'] = cuadro

    # ---- G real 21: 1.6. Datos grafico base 21 A5:A56 B-H ----
    ws = wb["1.6. Datos grafico base 21"]
    real_21 = []
    keys = [(2,'sal_priv'),(3,'sal_pub_nac'),(4,'sal_pub_prov'),
            (5,'jub_min'),(6,'jub_no_min'),(7,'no_reg'),(8,'auh')]
    for r in range(5, 57):
        fecha = ws.cell(r, 1).value
        if not isinstance(fecha, datetime): continue
        point = {"fecha": fecha.strftime("%Y-%m-%d"), "label": fecha.strftime("%b-%y")}
        for c, key in keys:
            point[key] = fmt_n(ws.cell(r, c).value)
        if any(point[k] is not None for k in [x[1] for x in keys]):
            real_21.append(point)
    data['real_21'] = real_21

    return data

# =====================================================================
#  TIPO DE CAMBIO (TCR bandas + Rofex)
# =====================================================================
def extract_tipo_cambio(status):
    import openpyxl
    from datetime import datetime, date, timedelta
    if not os.path.exists(EXCEL_PATHS["rofex"]):
        status.fail("Rofex", f"No se encontró: {EXCEL_PATHS['rofex']}")
        return None

    data = {}

    # ---- TCN series (TCR bandas - hoja TCN) ----
    # Si está bloqueado/sincronizando, mantener datos previos
    tcn = None
    if os.path.exists(EXCEL_PATHS["tcr_bandas"]):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATHS["tcr_bandas"], data_only=True)
            ws = wb["TCN"]
            tcn = []
            for r in range(3, ws.max_row + 1):
                fecha = ws.cell(r, 3).value  # C
                if not isinstance(fecha, (datetime, date)): continue
                point = {
                    "fecha": fecha.strftime("%Y-%m-%d"),
                    "oficial":          fmt_n(ws.cell(r, 4).value),    # D
                    "ccl":              fmt_n(ws.cell(r, 5).value),    # E
                    "banda_sup_previa": fmt_n(ws.cell(r, 6).value),    # F
                    "banda_sup_actual": fmt_n(ws.cell(r, 7).value),    # G
                }
                if any(v is not None for k, v in point.items() if k != 'fecha'):
                    tcn.append(point)
        except Exception as e:
            status.warn("TCR bandas", f"no se pudo leer ({e}); mantengo datos previos")
            tcn = None

    if tcn is None:
        # Fallback: usar JSON existente
        old = load_existing("tipo-cambio")
        if old and "tcn_series" in old:
            tcn = old["tcn_series"]
    data['tcn_series'] = tcn or []

    # ---- TCN serie LARGA desde com3500 (A3500 BCRA desde 2002) ----
    if os.path.exists(EXCEL_PATHS["com3500"]):
        try:
            import xlrd
            wb_c = xlrd.open_workbook(EXCEL_PATHS["com3500"])
            sh = wb_c.sheet_by_name("TCR diario y TCNPM")
            def xl_date(n):
                return (datetime(1899, 12, 30) + timedelta(days=int(n))).strftime("%Y-%m-%d")
            tcn_long = []
            for r in range(4, sh.nrows):
                fecha = sh.cell(r, 2).value
                val = sh.cell(r, 3).value
                if isinstance(fecha, (int, float)) and isinstance(val, (int, float)) and val > 0:
                    tcn_long.append({"fecha": xl_date(fecha), "oficial": val})
            data['tcn_long'] = tcn_long
            status.ok("Tipo de Cambio - com3500", f"{len(tcn_long)} puntos · {tcn_long[0]['fecha']} → {tcn_long[-1]['fecha']}")
        except ImportError:
            status.warn("Tipo de Cambio - com3500", "falta xlrd (instalar: pip install xlrd)")
        except Exception as e:
            status.warn("Tipo de Cambio - com3500", f"error: {e}")
    else:
        status.warn("Tipo de Cambio - com3500", "archivo no encontrado")

    # ---- Rofex (Cuadros para informes) ----
    # Salta filas y columnas OCULTAS del Excel (regla del usuario)
    wb2 = openpyxl.load_workbook(EXCEL_PATHS["rofex"], data_only=True)
    ws2 = wb2["Cuadros para informes"]

    # Detectar ocultas
    hidden_cols = set()
    for letter, dim in ws2.column_dimensions.items():
        if dim.hidden:
            hidden_cols.add(openpyxl.utils.column_index_from_string(letter))
    hidden_rows = set()
    for r_num, dim in ws2.row_dimensions.items():
        if dim.hidden:
            hidden_rows.add(r_num)

    BLOQUES = [
        {"id":"nominal", "label":"Nominal $/USD",  "fmt":"money", "header_row":5,  "month_row":6,  "data_start":7,  "data_end":29},
        {"id":"tna",     "label":"TNA %",          "fmt":"pct",   "header_row":36, "month_row":37, "data_start":38, "data_end":60},
        {"id":"varmens", "label":"Var. % mensual", "fmt":"pct",   "header_row":66, "month_row":67, "data_start":68, "data_end":90},
        {"id":"interes", "label":"Interés abierto","fmt":"int",   "header_row":96, "month_row":97, "data_start":98, "data_end":120},
    ]
    rofex = {}
    for b in BLOQUES:
        # Meses: saltar columnas ocultas
        meses = []
        col_indices = []
        for c in range(3, 25):
            if c in hidden_cols: continue
            v = ws2.cell(b["month_row"], c).value
            if isinstance(v, (datetime, date)):
                meses.append(v.strftime("%Y-%m"))
                col_indices.append(c)
            elif v is None and meses:
                break

        # Filas: saltar ocultas y filas sin fecha
        filas = []
        for r in range(b["data_start"], b["data_end"]+1):
            if r in hidden_rows: continue
            fecha = ws2.cell(r, 2).value
            if not isinstance(fecha, (datetime, date)): continue
            valores = []
            for c in col_indices:
                v = ws2.cell(r, c).value
                if isinstance(v, (int, float)):
                    valores.append(None if v == 0 else v)
                else:
                    valores.append(None)
            if any(v is not None for v in valores):
                filas.append({"fecha": fecha.strftime("%Y-%m-%d"), "valores": valores})

        # Drop columnas que están TODAS en None (meses vencidos sin datos)
        keep_idx = []
        for c_i in range(len(meses)):
            if any(row['valores'][c_i] is not None for row in filas):
                keep_idx.append(c_i)
        meses = [meses[i] for i in keep_idx]
        for row in filas:
            row['valores'] = [row['valores'][i] for i in keep_idx]

        rofex[b["id"]] = {"label": b["label"], "fmt": b["fmt"], "meses": meses, "filas": filas}
    data['rofex'] = rofex

    return data

# =====================================================================
#  MAIN
# =====================================================================
def main():
    print("=" * 64)
    print("ECO GO DASHBOARD - Refresh de datos")
    print("=" * 64)
    print(f"Dashboard dir: {DASHBOARD_DIR}")
    print(f"Excel base:    {BASE_EXCEL}")
    print(f"Data dir:      {DATA_DIR}")
    print()

    if not os.path.isdir(DATA_DIR):
        print(f"ERROR: no existe la carpeta {DATA_DIR}")
        return 1

    # Verificar openpyxl
    try:
        import openpyxl  # noqa
    except ImportError:
        print("\nERROR: falta el modulo 'openpyxl'.")
        print("Instalalo abriendo cmd y corriendo:")
        print("   pip install openpyxl")
        return 1

    status = Status()

    # ---- Precios ----
    print("[1/5] Procesando Precios...")
    try:
        d = extract_precios(status)
        if d:
            sz = save_data("precios", d)
            status.ok("Precios", f"{sz:,} bytes")
    except Exception as e:
        status.fail("Precios", str(e))
        traceback.print_exc()

    # ---- Empleo ----
    print("\n[2/5] Procesando Empleo...")
    try:
        d = extract_empleo(status)
        if d:
            sz = save_data("empleo", d)
            status.ok("Empleo", f"{sz:,} bytes")
    except Exception as e:
        status.fail("Empleo", str(e))
        traceback.print_exc()

    # ---- Salarios ----
    print("\n[3/5] Procesando Salarios...")
    try:
        d = extract_salarios(status)
        if d:
            sz = save_data("salarios", d)
            status.ok("Salarios", f"{sz:,} bytes")
    except Exception as e:
        status.fail("Salarios", str(e))
        traceback.print_exc()

    # ---- Tipo de Cambio ----
    print("\n[4/5] Procesando Tipo de Cambio (TCR bandas + Rofex)...")
    try:
        d = extract_tipo_cambio(status)
        if d:
            sz = save_data("tipo-cambio", d)
            status.ok("Tipo de Cambio", f"{sz:,} bytes · {len(d.get('tcn_series',[]))} pts TCN")
    except Exception as e:
        status.fail("Tipo de Cambio", str(e))
        traceback.print_exc()

    # ---- Internacional (Monitor mundial) ----
    print("\n[5/5] Procesando Internacional (Monitor mundial)...")
    try:
        if not os.path.exists(MONITOR_MUNDIAL_JS):
            status.warn("Internacional", f"no se encontró {MONITOR_MUNDIAL_JS}")
        else:
            with open(MONITOR_MUNDIAL_JS, encoding='utf-8') as f:
                src = f.read()
            m = re.search(r'window\.MONITOR_DATA\s*=\s*(\{.*\});?\s*$', src, re.DOTALL)
            if not m:
                status.fail("Internacional", "no pude parsear monitor-data.js")
            else:
                data = json.loads(m.group(1))
                # Reusar save_data pero con nombre 'internacional' (var INTERNACIONAL_DATA)
                json_path = os.path.join(DATA_DIR, "internacional.json")
                js_path   = os.path.join(DATA_DIR, "internacional.js")
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, default=str)
                js = f"// Datos Internacional - regenerado por refresh.py el {datetime.now().strftime('%Y-%m-%d %H:%M')}\nwindow.INTERNACIONAL_DATA = {json.dumps(data, ensure_ascii=False, default=str)};\n"
                with open(js_path, "w", encoding="utf-8") as f:
                    f.write(js)
                sz = os.path.getsize(json_path)
                status.ok("Internacional", f"{sz:,} bytes · {len(data.get('countries',[]))} países · {len(data.get('indicators',[]))} indicadores")
    except Exception as e:
        status.fail("Internacional", str(e))
        traceback.print_exc()

    # ---- Resumen ----
    print()
    print("=" * 64)
    ok    = sum(1 for r in status.results 