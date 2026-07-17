"""
ECO GO Dashboard - Refresh de datos
====================================
Re-lee los Excel originales y regenera los archivos JSON/JS del dashboard.

Uso:
  - Doble click a refresh.bat (recomendado)
  - O desde terminal: python refresh.py

Los archivos Excel NO se modifican (read-only).
"""

import os
import sys
import json
import re
import traceback
from datetime import datetime

# Asegurar consola con soporte unicode en Windows
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

# =====================================================================
#  CONFIGURACIÓN — Editá estas rutas si los Excel cambian de ubicación
# =====================================================================
# Directorio del dashboard = donde está este script
# (definido aquí para que BASE_EXCEL pueda derivarse de él en Linux)
DASHBOARD_DIR = os.path.dirname(os.path.abspath(__file__))

import platform as _platform
if _platform.system() == 'Windows':
    BASE_EXCEL = r"C:\Users\fscalise\OneDrive - ECOGO S.A\BD"
else:
    # En Linux (sandbox Cowork): el dashboard está en BD/07 Tableros/EcoGo-Dashboard
    # → BASE_EXCEL es dos niveles arriba
    BASE_EXCEL = os.path.dirname(os.path.dirname(DASHBOARD_DIR))

EXCEL_PATHS = {
    "ipc":         os.path.join(BASE_EXCEL, "Precios", "IPC TODESCA.xlsx"),
    "gd":          os.path.join(BASE_EXCEL, "Precios", "Gráficos de dispersión - copia - copia.xlsx"),
    "cm":          os.path.join(BASE_EXCEL, "Precios", "CM - DB.xlsx"),  # opcional (ya no se usa para proyeccion)
    "rpm_cuadro":  os.path.join(BASE_EXCEL, "Precios", "RPM Estudio Bein", "Alimentos Scrapping", "Cuadro Mensual - Capítulos Nuevo.xlsx"),
    "empleo":      os.path.join(BASE_EXCEL, "Empleo",  "Empleo_nuevo.xlsx"),
    "salarios":    os.path.join(BASE_EXCEL, "Empleo",  "Salarios.xlsx"),
    "tcr_bandas":  os.path.join(BASE_EXCEL, "Tipo de Cambio", "TCR bandas.xlsx"),
    "rofex":       os.path.join(BASE_EXCEL, "Tipo de Cambio", "Rofex.xlsx"),
    "com3500":     os.path.join(BASE_EXCEL, "Tipo de Cambio", "com3500.xls"),  # TCN A3500 historico
    "copia_blue":  os.path.join(BASE_EXCEL, "Tipo de Cambio", "Copia de Blue.xlsx"),  # Blue, MEP, CCL
    "base_esae":      os.path.join(BASE_EXCEL, "Actividad", "02 Indicador de Actividad CN2004", "Base EsAE.xlsx"),
    "emae_nuevo":     os.path.join(BASE_EXCEL, "Actividad", "EMAE - NUEVO.xlsx"),
    "pasivos_res":    os.path.join(BASE_EXCEL, "Monetarias", "pasivos reservas.xlsx"),
    "res_dep":        os.path.join(BASE_EXCEL, "Monetarias", "Reservas brutas y depósitos.xlsx"),
    "rigi":           os.path.join(BASE_EXCEL, "Códigos", "Python", "Patru", "rigi", "Proyectos RIGI.xlsx"),
}

# Monitor mundial — no es Excel, es un .js con datos del monitor externo
MONITOR_MUNDIAL_JS = os.path.join(BASE_EXCEL, "Internacional", "Monitor mundial", "data", "monitor-data.js")

DATA_DIR = os.path.join(DASHBOARD_DIR, "assets", "data")

# =====================================================================
#  Helpers
# =====================================================================


def _open_wb(path, data_only=True, read_only=True):
    """Abre un Excel leyendo primero a bytes para evitar problemas con OneDrive mount."""
    import io as _io
    import openpyxl as _opx
    with open(path, 'rb') as _f:
        _data = _f.read()
    return _opx.load_workbook(_io.BytesIO(_data), data_only=data_only, read_only=read_only)

def _read_text(path, encoding='utf-8'):
    """Lee un archivo de texto evitando [Errno 22] del mount de OneDrive."""
    import subprocess as _sub
    r = _sub.run(['cat', path], capture_output=True)
    if r.returncode == 0:
        return r.stdout.decode(encoding)
    # fallback directo
    with open(path, encoding=encoding) as f:
        return f.read()

class Status:
    def __init__(self):
        self.results = []
    def ok(self, name, detail=""):
        self.results.append(("OK", name, detail))
        print(f"  [OK]   {name}" + (f" — {detail}" if detail else ""))
    def warn(self, name, detail=""):
        self.results.append(("WARN", name, detail))
        print(f"  [---]  {name}" + (f" — {detail}" if detail else ""))
    def fail(self, name, detail=""):
        self.results.append(("FAIL", name, detail))
        print(f"  [FAIL] {name}" + (f" — {detail}" if detail else ""))

def save_data(name, data):
    """Guarda como JSON + JS con la variable global esperada."""
    json_path = os.path.join(DATA_DIR, f"{name}.json")
    js_path   = os.path.join(DATA_DIR, f"{name}.js")
    json_str  = json.dumps(data, ensure_ascii=False, default=str)
    try:
        with open(json_path, "w", encoding="utf-8") as f:
            f.write(json_str)
    except OSError:
        pass  # OneDrive bloquea .json existentes — solo escribimos el .js
    var_name = {"precios":"PRECIOS_DATA","empleo":"EMPLEO_DATA","salarios":"SALARIOS_DATA","tipo-cambio":"TC_DATA","reservas":"RESERVAS_DATA"}.get(name, name.upper()+"_DATA")
    js = f"// Datos de {name} - regenerado por refresh.py el {datetime.now().strftime('%Y-%m-%d %H:%M')}\nwindow.{var_name} = {json_str};\n"
    with open(js_path, "w", encoding="utf-8") as f:
        f.write(js)
    return os.path.getsize(js_path)

def load_existing(name):
    json_path = os.path.join(DATA_DIR, f"{name}.json")
    if os.path.exists(json_path):
        try:
            with open(json_path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None
    return None

def col_idx(letter):
    n = 0
    for c in letter:
        n = n * 26 + ord(c) - 64
    return n

def fmt_n(v):
    return v if isinstance(v, (int, float)) else None

# =====================================================================
#  PRECIOS
# =====================================================================
def extract_precios(status):
    import openpyxl

    if not os.path.exists(EXCEL_PATHS["ipc"]):
        status.fail("Precios", f"No se encontró: {EXCEL_PATHS['ipc']}")
        return None
    if not os.path.exists(EXCEL_PATHS["gd"]):
        status.fail("Precios - GD", f"No se encontró: {EXCEL_PATHS['gd']}")
        return None

    data = {}

    # ---- IPC último + series ----
    # read_only=False para acceso random por celda (col DK = 115) — BytesIO ya está en RAM
    wb = _open_wb(EXCEL_PATHS["ipc"], read_only=False)
    ws = wb["1. Nuevo IPC Nacional"]

    last = None
    for r in range(ws.max_row, 0, -1):
        if ws.cell(r, 20).value is not None:  # col T (nucleo m/m)
            last = r
            break

    def g(r, col):
        return ws.cell(r, col_idx(col)).value

    # Último mes
    fecha = g(last, 'A')
    data['ultimo'] = {
        "fecha": fecha.strftime("%Y-%m-%d") if isinstance(fecha, datetime) else str(fecha),
        "general_mm":   fmt_n(g(last, 'AF')),
        "general_ia":   fmt_n(g(last, 'AT')),
        "nucleo_mm":    fmt_n(g(last, 'T')),
        "nucleo_ia":    fmt_n(g(last, 'U')),
        "estacional_mm": fmt_n(g(last, 'P')),
        "estacional_ia": fmt_n(g(last, 'BV')),
        "regulados_mm": fmt_n(g(last, 'V')),
        "regulados_ia": fmt_n(g(last, 'W'))
    }

    # Series 24 meses
    series = []
    for r in range(max(6, last - 24), last + 1):
        fecha = g(r, 'A')
        if not isinstance(fecha, datetime): continue
        series.append({
            "fecha": fecha.strftime("%Y-%m-%d"),
            "general_mm": fmt_n(g(r, 'AF')), "general_ia": fmt_n(g(r, 'AT')),
            "nucleo_mm": fmt_n(g(r, 'T')),   "nucleo_ia": fmt_n(g(r, 'U')),
            "estacional_mm": fmt_n(g(r, 'P')), "estacional_ia": fmt_n(g(r, 'BV')),
            "regulados_mm": fmt_n(g(r, 'V')), "regulados_ia": fmt_n(g(r, 'W'))
        })
    data['series_24m'] = series

    # ---- Chart 21: '1. Nuevo IPC Nacional'!A17:A114, DJ, DK, AT ----
    def range_get(sheet, col, r1, r2):
        out = []
        for r in range(r1, r2+1):
            v = sheet.cell(r, col_idx(col)).value
            out.append(v)
        return out

    cats = range_get(ws, 'A', 17, 114)
    chart21 = []
    for i, c in enumerate(cats):
        if c is None: continue
        chart21.append({
            "fecha": c.strftime("%Y-%m-%d") if isinstance(c, datetime) else str(c),
            "mensual": range_get(ws,'DJ',17,114)[i],
            "promedio_anual": range_get(ws,'DK',17,114)[i],
            "var_ia": range_get(ws,'AT',17,114)[i]
        })
    data['chart21'] = chart21

    # ---- Chart 23: A6:A115, AF, AT, U, W, DK ----
    cats5 = range_get(ws, 'A', 6, 115)
    chart23 = []
    var_men = range_get(ws,'AF',6,115)
    var_ia5 = range_get(ws,'AT',6,115)
    nucleo_ia = range_get(ws,'U',6,115)
    w_series = range_get(ws,'W',6,115)
    prom_men = range_get(ws,'DK',6,115)
    for i, c in enumerate(cats5):
        if c is None: continue
        chart23.append({
            "fecha": c.strftime("%Y-%m-%d") if isinstance(c, datetime) else str(c),
            "var_men": var_men[i], "var_ia": var_ia5[i],
            "nucleo_ia": nucleo_ia[i], "regulados_ia": w_series[i],
            "prom_men": prom_men[i]
        })
    data['chart23'] = chart23

    # ---- Chart 22 (RPM): '4. Proyecciones'!B100:B201, C/E/F/H ----
    ws2 = wb["4. Proyecciones"]
    chart22 = []
    for r in range(100, 202):
        fecha = ws2.cell(r, 2).value
        if fecha is None: continue
        chart22.append({
            "fecha": fecha.strftime("%Y-%m-%d") if isinstance(fecha, datetime) else str(fecha),
            "rpm_mm": fmt_n(ws2.cell(r, 3).value),
            "ipc_gba": fmt_n(ws2.cell(r, 5).value),
            "ipc_nac": fmt_n(ws2.cell(r, 6).value),
            "rpm_ia": fmt_n(ws2.cell(r, 8).value)
        })
    data['chart22'] = chart22

    # ---- Chart 5: Gráficos de dispersión - Hoja3 ----
    wb_gd = _open_wb(EXCEL_PATHS["gd"])
    ws_gd = wb_gd["Hoja3"]
    CATS = {
        'Pan y cereales':'Bienes','Carnes y derivados':'Bienes',
        'Leche, productos lácteos y huevos':'Bienes','Aceites, grasas y manteca':'Bienes',
        'Frutas y verduras':'Bienes','Azúcar, dulces, chocolates':'Bienes',
        'Bebidas':'Bienes','Tabaco':'Bienes','Indumentaria':'Bienes',
        'Medicamentos':'Bienes','Medicamentos ':'Bienes',
        'Adquisición de vehículos':'Bienes','Combustibles':'Bienes',
        'Electrodomésticos':'Bienes','Bienes y servicios para la conservación del hogar':'Bienes',
        'Bienes':'Bienes',
        'Vivienda':'Servicios','Alquiler de la vivienda y gastos conexos':'Servicios',
        'Prepagas':'Servicios','Servicios recreativos':'Servicios','Servicios recreativos ':'Servicios',
        'Educación':'Servicios','Restaurantes':'Servicios','Servicios':'Servicios',
        'Electricidad, gas y otros combustibles':'Regulados','Transporte público':'Regulados',
        'Servicios  de telefonía e internet':'Regulados','Servicios de telefonía e internet':'Regulados',
        'Tarifa de Agua':'Regulados','Regulados':'Regulados',
        'Salarios Formales Privados':'Salarios','Salarios Informales':'Salarios',
        'Salarios Informales ':'Salarios','Jubilación Mínima c/Bono':'Salarios',
        'Jubilación s/Bono':'Salarios','Salarios Formales Públicos':'Salarios',
        'AUH':'Salarios','Salario de Cuentapropistas':'Salarios',
        'Dólar CCL':'Dólar','Dólar Oficial':'Dólar',
        'ICC(Materiales p/construcción)':'Construcción','ICC (Nivel Gral.)':'Construcción',
        'ICC (Mano de Obra)':'Construcción','ICC (Gastos Generales)':'Construcción',
        'IPIM (Nivel General)':'Mayorista','IPIM (Nacionales)':'Mayorista',
        'IPIM (Importados)':'Mayorista'
    }
    points = []
    for r in list(range(3, 42)) + list(range(44, 48)):
        label = ws_gd.cell(r, 1).value
        x = ws_gd.cell(r, 6).value
        y = ws_gd.cell(r, 7).value
        if label and x is not None and y is not None:
            cat = CATS.get(label, 'Otros') if not isinstance(label, str) else CATS.get(label.strip(), 'Otros')
            points.append({"label": str(label).strip(), "categoria": cat, "x": x, "y": y})
    def _fmt_scatter_label(raw):
        """'Nov23-Jun26' → 'Nov 2023 – Jun 2026 (gap vs Nivel general)'"""
        import re as _re
        if not raw:
            return str(raw or "")
        mt = _re.match(r'([A-Za-z]+)(\d{2})-([A-Za-z]+)(\d{2})', str(raw).strip())
        if mt:
            return f"{mt.group(1)} 20{mt.group(2)} – {mt.group(3)} 20{mt.group(4)} (gap vs Nivel general)"
        return str(raw)

    data['chart5'] = {
        "title": "Precios relativos: cuánto corregimos y cuánto falta corregir",
        "x_label": _fmt_scatter_label(ws_gd.cell(1, 6).value),
        "y_label": _fmt_scatter_label(ws_gd.cell(1, 7).value),
        "points": points
    }

    # ---- Proyección RPM (Cuadro Mensual - Capítulos Nuevo.xlsx, hoja siguiente a 'base') ----
    if os.path.exists(EXCEL_PATHS["rpm_cuadro"]):
        try:
            wb_cm = _open_wb(EXCEL_PATHS["rpm_cuadro"])
            # La hoja de datos es siempre la que está al lado de 'base' (índice 1)
            # Las hojas tienen formato YYMM (ej: '2605'), ordenadas de más reciente a más antigua
            idx_base = next((i for i, s in enumerate(wb_cm.sheetnames) if s.strip().lower() == 'base'), None)
            if idx_base is not None and idx_base + 1 < len(wb_cm.sheetnames):
                hoja = wb_cm.sheetnames[idx_base + 1]
            else:
                # Fallback: hoja YYMM más reciente
                yymm = sorted([s for s in wb_cm.sheetnames if s.strip().isdigit() and len(s.strip()) == 4], reverse=True)
                hoja = yymm[0] if yymm else None
            if hoja:
                ws_cm = wb_cm[hoja]
                # Sección limpia: filas 26-40 (1-indexed), cols A-D
                # Fila 26: título, fila 30+: datos (capítulo, mensual, anual, acumulada)
                rows = list(ws_cm.iter_rows(min_row=26, max_row=42, max_col=4, values_only=True))
                titulo = str(rows[0][0] or "").strip()
                proy = {
                    "titulo": titulo,
                    "header_periodo": f"Proyección — RPM Eco Go · {hoja}",
                    "filas": []
                }
                for r in rows:
                    cap = r[0]
                    if not cap or not isinstance(cap, str): continue
                    cap = cap.strip()
                    if not cap or cap.startswith("Fuente") or cap.startswith("PIEBGEB"): continue
                    mensual = r[1]; anual = r[2]; acum = r[3]
                    if mensual is None and anual is None: continue
                    proy["filas"].append({
                        "capitulo": cap,
                        "mensual": fmt_n(mensual),
                        "anual":   fmt_n(anual),
                        "acumulada": fmt_n(acum)
                    })
                data['proyeccion'] = proy
                status.ok("Precios - Proyección RPM", f"hoja {hoja}, {len(proy['filas'])} filas")
            else:
                _keep_old_proy(data, status)
        except Exception as e:
            status.warn("Precios - Proyección RPM", f"error: {e}; mantengo datos previos")
            _keep_old_proy(data, status)
    else:
        _keep_old_proy(data, status)

    return data

def _keep_old_proy(data, status):
    """Si no hay CM-DB, conservar la proyección anterior del JSON existente."""
    old = load_existing("precios")
    if old and "proyeccion" in old:
        data["proyeccion"] = old["proyeccion"]
        status.warn("Precios - Proyección RPM", "CM-DB no encontrado, mantengo dato previo")
    else:
        status.warn("Precios - Proyección RPM", "sin CM-DB ni dato previo")

# =====================================================================
#  EMPLEO (EPH + Cuadro Trim + SIPA + Provincias)
# =====================================================================
def extract_empleo(status):
    import openpyxl
    if not os.path.exists(EXCEL_PATHS["empleo"]):
        status.fail("Empleo", f"No se encontró: {EXCEL_PATHS['empleo']}")
        return None

    wb = _open_wb(EXCEL_PATHS["empleo"])
    data = {}

    # ---- Tasas EPH (sección principal R8-R124, cols B,E,H,K,N,Q,T) ----
    ws = wb["Tasas EPH"]
    eph = []
    for r in range(8, 125):
        fecha = ws.cell(r, 1).value
        if fecha is None: continue
        if isinstance(fecha, datetime):
            fecha_str = fecha.strftime("%Y-%m-%d")
            ord_idx = (fecha.year * 4) + (fecha.month - 1) // 3
            label = fecha.strftime("%b-%Y")
        elif isinstance(fecha, str):
            m = re.match(r'(IV|III|II|I)-(\d{2,4})', fecha.strip())
            if m:
                qmap = {'I':1,'II':2,'III':3,'IV':4}
                q = qmap[m.group(1)]
                yy = int(m.group(2))
                if yy < 100: yy += 2000
                label = fecha.strip()
            else:
                m2 = re.match(r'(\d)\s*º?\s*trim\s*(\d{4})', fecha.strip(), re.IGNORECASE)
                if not m2: continue
                q = int(m2.group(1))
                yy = int(m2.group(2))
                label = f"{['','I','II','III','IV'][q]}-{str(yy)[2:]}"
            fecha_str = f"{yy}-{(q-1)*3+1:02d}-01"
            ord_idx = (yy * 4) + (q - 1)
        else:
            continue
        eph.append({
            "fecha": fecha_str, "ord": ord_idx, "label": label,
            "actividad":     fmt_n(ws.cell(r, 2).value),
            "empleo":        fmt_n(ws.cell(r, 5).value),
            "desocup":       fmt_n(ws.cell(r, 8).value),
            "ocup_dem":      fmt_n(ws.cell(r, 11).value),
            "subocup":       fmt_n(ws.cell(r, 14).value),
            "subocup_dem":   fmt_n(ws.cell(r, 17).value),
            "subocup_nodem": fmt_n(ws.cell(r, 20).value)
        })
    eph.sort(key=lambda x: x['ord'])
    data['eph'] = eph
    data['eph_ultimo'] = eph[-1] if eph else None

    # ---- Cuadro empleo trim ----
    ws = wb["Cuadro empleo trim"]
    COLS_TRIM = [(3,'2012 (est.)'), (4,'II-18'), (5,'II-23'), (7,'IV-23'), (8,'I-24'),
                 (9,'II-24'), (10,'IV-24'), (12,'I-25'), (13,'III-25'), (14,'IV-25')]
    trim = {"periodos": [c[1] for c in COLS_TRIM], "filas": []}
    for r in range(5, 20):
        cat = ws.cell(r, 2).value
        if not cat: continue
        valores = []
        for col_idx_n, _ in COLS_TRIM:
            v = ws.cell(r, col_idx_n).value
            valores.append(v if isinstance(v, (int, float)) else None)
        trim["filas"].append({"categoria": str(cat).strip(), "valores": valores})
    data['trim'] = trim

    # ---- Cuadro SIPA ext (2) ----
    ws = wb["Cuadro SIPA ext (2)"]
    COLS_SIPA = [
        (3,'Jun-12','en miles','stock'),(4,'Ago-18*','en miles','stock'),
        (5,'Dic-23*','en miles','stock'),(6,'Feb-25*','en miles','stock'),
        (9,'Ene-26*','en miles','stock'),(12,'Feb-26*','en miles','stock'),
        (13,'Dif. (6)-(1)','vs Jun-12','diff'),(14,'Dif. (6)-(2)','vs Ago-18','diff'),
        (15,'Dif. (6)-(3)','vs Dic-23','diff'),(16,'Dif. (6)-(4)','vs Feb-25','diff'),
        (17,'Dif. (6)-(5)','vs Ene-26','diff')
    ]
    sipa = {"cols": [{"header":c[1],"sub":c[2],"tipo":c[3]} for c in COLS_SIPA], "filas":[]}
    for r in range(4, 13):
        cat = ws.cell(r, 2).value
        if not cat: continue
        valores = []
        for col_idx_n, _, _, _ in COLS_SIPA:
            v = ws.cell(r, col_idx_n).value
            valores.append(v if isinstance(v, (int, float)) else None)
        sipa["filas"].append({"categoria": str(cat).strip(), "valores": valores})
    data['sipa'] = sipa

    # ---- Provincias (Hoja7) ----
    ws = wb["Hoja7"]
    provincias = {"fechas": [], "datos": []}
    for c in range(3, 9):
        v = ws.cell(3, c).value
        if isinstance(v, datetime):
            provincias["fechas"].append(v.strftime("%Y-%m"))
    for r in range(5, 30):
        prov = ws.cell(r, 2).value
        if not prov: continue
        vals = []
        for c in range(3, 9):
            v = ws.cell(r, c).value
            vals.append(v if isinstance(v, (int, float)) else None)
        if any(v is not None for v in vals):
            provincias["datos"].append({"provincia": str(prov).strip(), "valores": vals})
    data['provincias'] = provincias

    return data

# =====================================================================
#  SALARIOS
# =====================================================================
def extract_salarios(status):
    import openpyxl
    if not os.path.exists(EXCEL_PATHS["salarios"]):
        status.fail("Salarios", f"No se encontró: {EXCEL_PATHS['salarios']}")
        return None

    wb = _open_wb(EXCEL_PATHS["salarios"])
    data = {}

    # ---- G Sal real: 1.1 INDEC A186:A298, CE/CF/CH/CI ----
    ws = wb["1.1 INDEC"]
    sal_real = []
    for r in range(186, 299):
        fecha = ws.cell(r, 1).value
        if not isinstance(fecha, datetime): continue
        priv = fmt_n(ws.cell(r, 83).value)  # CE
        pub  = fmt_n(ws.cell(r, 84).value)  # CF
        nor  = fmt_n(ws.cell(r, 86).value)  # CH
        tot  = fmt_n(ws.cell(r, 87).value)  # CI
        if any(v is not None for v in [priv, pub, nor, tot]):
            sal_real.append({
                "fecha": fecha.strftime("%Y-%m-%d"),
                "label": fecha.strftime("%b-%y"),
                "priv": priv, "pub": pub, "nor": nor, "tot": tot
            })
    data['sal_real'] = sal_real

    # ---- Cuadro INDEC 1.3 ----
    ws = wb["1.3 Cuadro INDEC"]
    fecha_actual = ws.cell(3, 2).value
    fecha_compar = ws.cell(3, 11).value
    def iso(v):
        if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
        return str(v) if v else ''
    cuadro = {
        "fecha_actual": iso(fecha_actual),
        "fecha_comparacion": iso(fecha_compar),
        "headers": [],
        "filas": []
    }
    for c in range(3, 11):
        v = ws.cell(4, c).value
        cuadro["headers"].append(str(v).replace('\n',' ').strip() if v else '')
    for r in range(5, 10):
        cat = ws.cell(r, 2).value
        if not cat: continue
        valores = []
        for c in range(3, 11):
            v = ws.cell(r, c).value
            valores.append(v if isinstance(v, (int, float)) else None)
        cuadro["filas"].append({"categoria": str(cat).strip(), "valores": valores})
    data['cuadro'] = cuadro

    # ---- G real 21: 1.6. Datos grafico base 21 A5:A56 B-H ----
    ws = wb["1.6. Datos grafico base 21"]
    real_21 = []
    keys = [(2,'sal_priv'),(3,'sal_pub_nac'),(4,'sal_pub_prov'),
            (5,'jub_min'),(6,'jub_no_min'),(7,'no_reg'),(8,'auh')]
    for r in range(5, 57):
        fecha = ws.cell(r, 1).value
        if not isinstance(fecha, datetime): continue
        point = {"fecha": fecha.strftime("%Y-%m-%d"), "label": fecha.strftime("%b-%y")}
        for c, key in keys:
            point[key] = fmt_n(ws.cell(r, c).value)
        if any(point[k] is not None for k in [x[1] for x in keys]):
            real_21.append(point)
    data['real_21'] = real_21

    return data

# =====================================================================
#  TIPO DE CAMBIO (TCR bandas + Rofex)
# =====================================================================
def extract_tipo_cambio(status):
    import openpyxl
    from datetime import datetime, date, timedelta
    if not os.path.exists(EXCEL_PATHS["rofex"]):
        status.fail("Rofex", f"No se encontró: {EXCEL_PATHS['rofex']}")
        return None

    data = {}

    # ---- TCN series (TCR bandas - hoja TCN) ----
    # Col C=fecha, D=oficial, E=CCL, F=Banda inferior, G=Banda superior
    # Si está bloqueado/sincronizando, mantener datos previos
    tcn = None
    if os.path.exists(EXCEL_PATHS["tcr_bandas"]):
        try:
            wb = _open_wb(EXCEL_PATHS["tcr_bandas"])
            ws = wb["TCN"]
            tcn = []
            for row in ws.iter_rows(min_row=3, values_only=True):  # iter_rows es O(n), no O(n²)
                fecha = row[2]  # col C = índice 2
                if not isinstance(fecha, (datetime, date)): continue
                point = {
                    "fecha":    fecha.strftime("%Y-%m-%d"),
                    "oficial":  fmt_n(row[3]),   # D
                    "ccl":      fmt_n(row[4]),   # E
                    "banda_inf":fmt_n(row[5]),   # F
                    "banda_sup":fmt_n(row[6]),   # G
                }
                if any(v is not None for k, v in point.items() if k != 'fecha'):
                    tcn.append(point)
        except Exception as e:
            status.warn("TCR bandas", f"no se pudo leer ({e}); mantengo datos previos")
            tcn = None

    if tcn is None:
        # Fallback: usar JSON existente
        old = load_existing("tipo-cambio")
        if old and "tcn_series" in old:
            tcn = old["tcn_series"]
    data['tcn_series'] = tcn or []

    # ---- TCN serie LARGA desde com3500 (A3500 BCRA desde 2002) ----
    if os.path.exists(EXCEL_PATHS["com3500"]):
        try:
            import xlrd
            wb_c = xlrd.open_workbook(EXCEL_PATHS["com3500"])
            sh = wb_c.sheet_by_name("TCR diario y TCNPM")
            def xl_date(n):
                return (datetime(1899, 12, 30) + timedelta(days=int(n))).strftime("%Y-%m-%d")
            tcn_long = []
            for r in range(4, sh.nrows):
                fecha = sh.cell(r, 2).value
                val = sh.cell(r, 3).value
                if isinstance(fecha, (int, float)) and isinstance(val, (int, float)) and val > 0:
                    tcn_long.append({"fecha": xl_date(fecha), "oficial": val})
            data['tcn_long'] = tcn_long
            status.ok("Tipo de Cambio - com3500", f"{len(tcn_long)} puntos · {tcn_long[0]['fecha']} → {tcn_long[-1]['fecha']}")
        except ImportError:
            status.warn("Tipo de Cambio - com3500", "falta xlrd (instalar: pip install xlrd)")
        except Exception as e:
            status.warn("Tipo de Cambio - com3500", f"error: {e}")
    else:
        status.warn("Tipo de Cambio - com3500", "archivo no encontrado")

    # ---- Series Blue / MEP / CCL históricas (Copia de Blue.xlsx - hoja Diario) ----
    # Col A=fecha, B=Blue, C=MEP, D=CCL
    if os.path.exists(EXCEL_PATHS["copia_blue"]):
        try:
            wb_b = _open_wb(EXCEL_PATHS["copia_blue"])
            ws_b = wb_b["Diario"]
            blue_long, mep_long, ccl_long = [], [], []
            for row in ws_b.iter_rows(min_row=2, values_only=True):  # iter_rows es O(n), no O(n²)
                fecha = row[0]
                if not isinstance(fecha, (datetime, date)): continue
                fecha_str = fecha.strftime("%Y-%m-%d") if isinstance(fecha, datetime) else fecha.isoformat()
                blue_v = row[1]
                mep_v  = row[2]
                ccl_v  = row[3]
                if isinstance(blue_v, (int, float)) and blue_v > 0:
                    blue_long.append({"fecha": fecha_str, "valor": blue_v})
                if isinstance(mep_v, (int, float)) and mep_v > 0:
                    mep_long.append({"fecha": fecha_str, "valor": mep_v})
                if isinstance(ccl_v, (int, float)) and ccl_v > 0:
                    ccl_long.append({"fecha": fecha_str, "valor": ccl_v})
            data['blue_long'] = blue_long
            data['mep_long']  = mep_long
            data['ccl_long']  = ccl_long
            status.ok("Tipo de Cambio - Blue/MEP/CCL",
                      f"Blue {len(blue_long)} pts · MEP {len(mep_long)} pts · CCL {len(ccl_long)} pts")
        except Exception as e:
            status.warn("Tipo de Cambio - Blue/MEP/CCL", f"error: {e}")
    else:
        status.warn("Tipo de Cambio - Blue/MEP/CCL", "Copia de Blue.xlsx no encontrado")

    # ---- Rofex (Cuadros para informes) ----
    # Salta filas y columnas OCULTAS del Excel (regla del usuario)
    # read_only=False es necesario para que column_dimensions / row_dimensions funcionen
    wb2 = _open_wb(EXCEL_PATHS["rofex"], read_only=False)
    ws2 = wb2["Cuadros para informes"]

    # Detectar ocultas
    hidden_cols = set()
    for letter, dim in ws2.column_dimensions.items():
        if dim.hidden:
            hidden_cols.add(openpyxl.utils.column_index_from_string(letter))
    hidden_rows = set()
    for r_num, dim in ws2.row_dimensions.items():
        if dim.hidden:
            hidden_rows.add(r_num)

    BLOQUES = [
        {"id":"nominal", "label":"Nominal $/USD",  "fmt":"money", "header_row":5,  "month_row":6,  "data_start":7,  "data_end":29},
        {"id":"tna",     "label":"TNA %",          "fmt":"pct",   "header_row":36, "month_row":37, "data_start":38, "data_end":60},
        {"id":"varmens", "label":"Var. % mensual", "fmt":"pct",   "header_row":66, "month_row":67, "data_start":68, "data_end":90},
        {"id":"interes", "label":"Interés abierto","fmt":"int",   "header_row":96, "month_row":97, "data_start":98, "data_end":120},
    ]
    rofex = {}
    for b in BLOQUES:
        # Meses: saltar columnas ocultas
        meses = []
        col_indices = []
        for c in range(3, 25):
            if c in hidden_cols: continue
            v = ws2.cell(b["month_row"], c).value
            if isinstance(v, (datetime, date)):
                meses.append(v.strftime("%Y-%m"))
                col_indices.append(c)
            elif v is None and meses:
                break

        # Filas: saltar ocultas y filas sin fecha
        filas = []
        for r in range(b["data_start"], b["data_end"]+1):
            if r in hidden_rows: continue
            fecha = ws2.cell(r, 2).value
            if not isinstance(fecha, (datetime, date)): continue
            valores = []
            for c in col_indices:
                v = ws2.cell(r, c).value
                if isinstance(v, (int, float)):
                    valores.append(None if v == 0 else v)
                else:
                    valores.append(None)
            if any(v is not None for v in valores):
                filas.append({"fecha": fecha.strftime("%Y-%m-%d"), "valores": valores})

        # Drop columnas que están TODAS en None (meses vencidos sin datos)
        keep_idx = []
        for c_i in range(len(meses)):
            if any(row['valores'][c_i] is not None for row in filas):
                keep_idx.append(c_i)
        meses = [meses[i] for i in keep_idx]
        for row in filas:
            row['valores'] = [row['valores'][i] for i in keep_idx]

        rofex[b["id"]] = {"label": b["label"], "fmt": b["fmt"], "meses": meses, "filas": filas}
    data['rofex'] = rofex

    return data

# =====================================================================
#  EMAE SERIES (actividad.html → window.EMAE_SERIES)
# =====================================================================
def extract_emae_series(status):
    """
    Lee Base EsAE.xlsx y genera emae_series.js con la estructura:
      { original: [{date, original, desest}],
        sector_ce: [{date, vals:[16]}],
        sector_se: [{date, vals:[16]}],
        sector_labels: [16 strings] }

    Estructura esperada en Base EsAE.xlsx:
      - Hoja "EMAE"   (o similar): col A=fecha, B=original, C=desest
      - Hoja "Sector CE": col A=fecha, cols B-Q = 16 sectores (con estacionalidad)
      - Hoja "Sector SE": col A=fecha, cols B-Q = 16 sectores (sin estacionalidad)
      - Hoja "Labels" o fila 1 de Sector CE: nombres de los 16 sectores

    Si la hoja exacta no coincide, el código busca por nombre parcial.
    Ajustá SHEET_MAP abajo si los nombres reales difieren.
    """
    import io as _io
    import openpyxl as _opx
    from datetime import datetime as _dt

    SECTOR_LABELS = [
        'Agro', 'Pesca', 'Minería', 'Industria', 'Electricidad',
        'Construcción', 'Comercio', 'Hot. y rest.', 'Transp. y com.',
        'Int. financiera', 'Act. empresariales', 'Adm. pública',
        'Enseñanza', 'Serv. sociales', 'Otras act.', 'Imp. netos'
    ]

    path = EXCEL_PATHS["base_esae"]
    if not os.path.exists(path):
        status.warn("EMAE Series", f"no se encontró {path}")
        return None

    try:
        with open(path, 'rb') as _f:
            _data = _f.read()
        wb = _opx.load_workbook(_io.BytesIO(_data), data_only=True, read_only=True)
    except Exception as e:
        status.warn("EMAE Series", f"no se pudo abrir el Excel: {e}")
        return None

    def _find_sheet(wb, candidates):
        names_lower = {s.lower().strip(): s for s in wb.sheetnames}
        for c in candidates:
            if c.lower() in names_lower:
                return wb[names_lower[c.lower()]]
        # Búsqueda parcial
        for c in candidates:
            for k, v in names_lower.items():
                if c.lower() in k:
                    return wb[v]
        return None

    def _parse_date(v):
        if v is None: return None
        if isinstance(v, str):
            v = v.strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y"):
                try: return _dt.strptime(v, fmt)
                except: pass
        if hasattr(v, 'year'): return v
        return None

    def _fmt_date(v):
        if v is None: return None
        try: return v.strftime("%Y-%m")
        except: return str(v)[:7]

    # ---- Serie original + desestacionalizada ----
    ws_orig = _find_sheet(wb, ['EMAE', 'Original', 'Serie', 'Mensual', 'emae'])
    original = []
    if ws_orig:
        for row in ws_orig.iter_rows(min_row=2, max_col=3, values_only=True):
            d = _parse_date(row[0])
            if d is None: continue
            orig_val = fmt_n(row[1])
            dest_val = fmt_n(row[2])
            if orig_val is None and dest_val is None: continue
            original.append({"date": _fmt_date(d), "original": orig_val, "desest": dest_val})
    else:
        status.warn("EMAE Series", "no se encontró hoja de serie original (buscando: EMAE/Original/Serie/Mensual)")

    # ---- Sectores con estacionalidad ----
    ws_ce = _find_sheet(wb, ['CE', 'Sector CE', 'Con estacionalidad', 'sector_ce', 'sectores ce'])
    sector_ce = []
    if ws_ce:
        rows_ce = list(ws_ce.iter_rows(min_row=2, max_col=17, values_only=True))
        # Fila 1 puede tener labels — intentar leerlos
        hdr = list(ws_ce.iter_rows(min_row=1, max_row=1, max_col=17, values_only=True))
        if hdr and any(isinstance(v, str) for v in hdr[0][1:]):
            SECTOR_LABELS[:] = [str(v).strip() for v in hdr[0][1:17] if v is not None]
        for row in rows_ce:
            d = _parse_date(row[0])
            if d is None: continue
            vals = [fmt_n(row[i]) for i in range(1, 17)]
            if all(v is None for v in vals): continue
            sector_ce.append({"date": _fmt_date(d), "vals": vals})
    else:
        status.warn("EMAE Series", "no se encontró hoja de sectores CE")

    # ---- Sectores sin estacionalidad ----
    ws_se = _find_sheet(wb, ['SE', 'Sector SE', 'Sin estacionalidad', 'sector_se', 'sectores se'])
    sector_se = []
    if ws_se:
        for row in ws_se.iter_rows(min_row=2, max_col=17, values_only=True):
            d = _parse_date(row[0])
            if d is None: continue
            vals = [fmt_n(row[i]) for i in range(1, 17)]
            if all(v is None for v in vals): continue
            sector_se.append({"date": _fmt_date(d), "vals": vals})
    else:
        status.warn("EMAE Series", "no se encontró hoja de sectores SE")

    if not original and not sector_ce and not sector_se:
        status.fail("EMAE Series", f"no se extrajo ningún dato de {os.path.basename(path)}. Verificar nombres de hojas: {wb.sheetnames[:8]}")
        return None

    result = {
        "original":      original,
        "sector_ce":     sector_ce,
        "sector_se":     sector_se,
        "sector_labels": SECTOR_LABELS,
    }
    status.ok("EMAE Series", f"{len(original)} meses - {len(sector_ce)} CE - {len(sector_se)} SE")
    return result

def save_emae_series(data):
    """Guarda emae_series.js (sin .json ya que es solo lectura del dashboard)."""
    js_path = os.path.join(DATA_DIR, "emae_series.js")
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    js = (
        f"// EMAE series data - EcoGo - {ts}\n"
        f"window.EMAE_SERIES = {json.dumps(data, ensure_ascii=False, default=str)};\n"
    )
    with open(js_path, "w", encoding="utf-8") as f:
        f.write(js)
    return os.path.getsize(js_path)

# =====================================================================
#  RESERVAS (pasivos reservas.xlsx + Reservas brutas y depositos.xlsx)
# =====================================================================
def extract_reservas(status):
    import io as _io
    import openpyxl as _opx
    from datetime import datetime as _dt

    ROW_META = {
        'reservas brutas':        (0, False, None,   'Reservas brutas'),
        'inconvertible':          (1, False, None,   'Inconvertible'),
        'reservas convertibles':  (1, True,  'conv', 'Reservas convertibles'),
        'valores':                (2, False, 'conv', 'Valores'),
        'divisas':                (2, False, 'conv', 'Divisas'),
        'degs':                   (2, False, 'conv', 'DEGs'),
        'oro':                    (2, False, 'conv', 'Oro'),
        'otros':                  (2, False, 'conv', 'Otros'),
        'pasivos brutos':         (0, True,  'pas',  'Pasivos brutos'),
        'encajes':                (1, False, 'pas',  'Encajes'),
        'swap china':             (1, False, 'pas',  'Swap China'),
        'swap basilea':           (1, False, 'pas',  'Swap Basilea + CAF'),
        'swap sedesa':            (1, False, 'pas',  'Swap SEDESA / Repo'),
        'swap activado eeuu':     (1, False, 'pas',  'Swap activado EEUU'),
        'otras obligaciones':     (1, False, 'pas',  'Otras obligaciones'),
        'bopreal':                (1, False, 'pas',  'BOPREAL'),
        'depositos del gobierno': (1, False, 'pas',  'Depositos del Gobierno'),
        'reservas liquidas':      (0, False, None,   'Reservas liquidas'),
        'reservas netas':         (0, False, None,   'Reservas netas'),
        'rin gob':                (0, False, None,   'RIN Gob. Nac.'),
    }

    def _norm(s):
        import unicodedata as _u
        s = _u.normalize('NFD', str(s).lower())
        s = ''.join(c for c in s if _u.category(c) != 'Mn')
        return re.sub(r'[^a-z\s]', '', s).strip()

    def _match_meta(label):
        nl = _norm(label)
        for key, meta in ROW_META.items():
            if nl.startswith(key):
                return meta
        return None

    MESES = ['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic']
    def _fmt_date(dt):
        return f"{dt.day:02d}/{MESES[dt.month-1]}-{str(dt.year)[2:]}"

    data = {}

    # RIN cuadro
    rin_path = EXCEL_PATHS["pasivos_res"]
    if not os.path.exists(rin_path):
        status.warn("Reservas - RIN", f"no se encontro {rin_path}")
    else:
        try:
            with open(rin_path, 'rb') as _f: _d = _f.read()
            wb = _opx.load_workbook(_io.BytesIO(_d), data_only=True, read_only=True)
            ws = wb['Cuadro RIN']

            date_row = list(ws.iter_rows(min_row=4, max_row=4, max_col=60, values_only=True))[0]
            date_cols = []
            last_dt = None
            for ci, v in enumerate(date_row):
                if isinstance(v, _dt):
                    if last_dt is None or v > last_dt:
                        date_cols.append((ci, v, _fmt_date(v)))
                        last_dt = v

            dates_labels = [lbl for _, _, lbl in date_cols]
            col_indices  = [ci for ci, _, _ in date_cols]

            def _extract_vals(row):
                vals = []
                for ci in col_indices:
                    v = row[ci] if ci < len(row) else None
                    if isinstance(v, str) and v.startswith('#'): v = None
                    vals.append(round(v, 1) if isinstance(v, float) else v)
                return vals

            rows_out = []
            for row in ws.iter_rows(min_row=5, max_row=26, max_col=60, values_only=True):
                label = row[0]
                if not label or not isinstance(label, str): continue
                label = label.strip()
                if not label or label.startswith(('1 ', '2 ', '3 ')): continue
                meta = _match_meta(label)
                if meta is None: continue
                indent, expandable, group, display = meta
                rows_out.append({
                    'label': display, 'indent': indent,
                    'expandable': expandable, 'group': group,
                    'vals': _extract_vals(row),
                })

            for row in ws.iter_rows(min_row=27, max_row=36, max_col=60, values_only=True):
                label = row[0]
                if not label or not isinstance(label, str): continue
                if 'tmu' in _norm(label):
                    rows_out.append({
                        'label': 'RIN TMU FMI', 'indent': 0,
                        'expandable': False, 'group': None,
                        'vals': _extract_vals(row),
                    })
                    break

            data['rin'] = {'dates': dates_labels, 'rows': rows_out}
            status.ok("Reservas - RIN", f"{len(dates_labels)} fechas - {len(rows_out)} filas")
        except Exception as e:
            status.fail("Reservas - RIN", str(e))
            traceback.print_exc()

    # G5: serie diaria
    dep_path = EXCEL_PATHS["res_dep"]
    if not os.path.exists(dep_path):
        status.warn("Reservas - G5", f"no se encontro {dep_path}")
    else:
        try:
            with open(dep_path, 'rb') as _f: _d = _f.read()
            wb2 = _opx.load_workbook(_io.BytesIO(_d), data_only=True, read_only=True)
            ws2 = wb2['Datos']
            g5 = []
            for row in ws2.iter_rows(min_row=3, max_col=5, values_only=True):
                dt = row[0] or row[1]
                r   = fmt_n(row[2])
                dep = fmt_n(row[3])
                pre = fmt_n(row[4])
                if not isinstance(dt, _dt) or r is None: continue
                g5.append({'d': dt.strftime('%Y-%m-%d'), 'r': r, 'dep': dep, 'pre': pre})
            data['g5'] = g5
            status.ok("Reservas - G5", f"{len(g5)} dias - ultimo: {g5[-1]['d'] if g5 else '?'}")
        except Exception as e:
            status.fail("Reservas - G5", str(e))
            traceback.print_exc()

    return data if data else None

# =====================================================================
#  MAIN
# =====================================================================
def main():
    print("=" * 64)
    print("ECO GO DASHBOARD - Refresh de datos")
    print("=" * 64)
    print(f"Dashboard dir: {DASHBOARD_DIR}")
    print(f"Excel base:    {BASE_EXCEL}")
    print(f"Data dir:      {DATA_DIR}")
    print()

    if not os.path.isdir(DATA_DIR):
        print(f"ERROR: no existe la carpeta {DATA_DIR}")
        return 1

    try:
        import openpyxl  # noqa
    except ImportError:
        print("\nERROR: falta el modulo 'openpyxl'.")
        print("   pip install openpyxl")
        return 1

    status = Status()

    # ---- Precios ----
    print("[1/8] Procesando Precios...")
    try:
        d = extract_precios(status)
        if d:
            sz = save_data("precios", d)
            status.ok("Precios", f"{sz:,} bytes")
    except Exception as e:
        status.fail("Precios", str(e))
        traceback.print_exc()

    # ---- EMAE Series ----
    print("\n[2/8] Procesando EMAE Series (actividad)...")
    try:
        d = extract_emae_series(status)
        if d:
            sz = save_emae_series(d)
            status.ok("EMAE Series", f"{sz:,} bytes guardados en emae_series.js")
    except Exception as e:
        status.fail("EMAE Series", str(e))
        traceback.print_exc()

    # ---- Empleo ----
    print("\n[3/8] Procesando Empleo...")
    try:
        d = extract_empleo(status)
        if d:
            sz = save_data("empleo", d)
            status.ok("Empleo", f"{sz:,} bytes")
    except Exception as e:
        status.fail("Empleo", str(e))
        traceback.print_exc()

    # ---- Salarios ----
    print("\n[4/8] Procesando Salarios...")
    try:
        d = extract_salarios(status)
        if d:
            sz = save_data("salarios", d)
            status.ok("Salarios", f"{sz:,} bytes")
    except Exception as e:
        status.fail("Salarios", str(e))
        traceback.print_exc()

    # ---- Tipo de Cambio ----
    print("\n[5/8] Procesando Tipo de Cambio...")
    try:
        d = extract_tipo_cambio(status)
        if d:
            sz = save_data("tipo-cambio", d)
            status.ok("Tipo de Cambio", f"{sz:,} bytes - {len(d.get('tcn_series',[]))} pts TCN")
    except Exception as e:
        status.fail("Tipo de Cambio", str(e))
        traceback.print_exc()

    # ---- Reservas ----
    print("\n[6/8] Procesando Reservas...")
    try:
        d = extract_reservas(status)
        if d:
            sz = save_data("reservas", d)
            status.ok("Reservas", f"{sz:,} bytes")
    except Exception as e:
        status.fail("Reservas", str(e))
        traceback.print_exc()

    # ---- Internacional ----
    print("\n[7/8] Procesando Internacional (Monitor mundial)...")
    try:
        if not os.path.exists(MONITOR_MUNDIAL_JS):
            status.warn("Internacional", f"no se encontro {MONITOR_MUNDIAL_JS}")
        else:
            src = _read_text(MONITOR_MUNDIAL_JS)
            m = re.search(r'window\.MONITOR_DATA\s*=\s*(\{.*\});?\s*$', src, re.DOTALL)
            if not m:
                status.fail("Internacional", "no pude parsear monitor-data.js")
            else:
                intl_data = json.loads(m.group(1))
                json_path = os.path.join(DATA_DIR, "internacional.json")
                js_path   = os.path.join(DATA_DIR, "internacional.js")
                _intl_str = json.dumps(intl_data, ensure_ascii=False, default=str)
                try:
                    with open(json_path, "w", encoding="utf-8") as f:
                        f.write(_intl_str)
                except OSError:
                    pass
                js = f"// Datos Internacional - regenerado por refresh.py el {datetime.now().strftime('%Y-%m-%d %H:%M')}\nwindow.INTERNACIONAL_DATA = {_intl_str};\n"
                with open(js_path, "w", encoding="utf-8") as f:
                    f.write(js)
                sz = os.path.getsize(js_path)
                status.ok("Internacional", f"{sz:,} bytes - {len(intl_data.get('countries',[]))} paises")
    except Exception as e:
        status.fail("Internacional", str(e))
        traceback.print_exc()

    # ---- Mercados (API) ----
    print("\n[8/8] Actualizando Mercados (EcoGo Markets API)...")
    try:
        import urllib.request
        MERCADOS_API = "https://ecogomarkets.honorio-zabaleta.workers.dev/data/dashboard_payload.json"
        req = urllib.request.Request(MERCADOS_API, headers={"User-Agent": "EcoGo-Dashboard-Refresh/1.0"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        subset = {
            "meta":                      payload.get("meta", {}),
            "external_context":          payload.get("external_context", {}),
            "overview":                  payload.get("overview", {}),
            "fixed_curve":               payload.get("fixed_curve", []),
            "fixed_curve_history":       payload.get("fixed_curve_history", {"dates": [], "curves": {}}),
            "cer_curve":                 payload.get("cer_curve", []),
            "cer_curve_history":         payload.get("cer_curve_history", {"dates": [], "curves": {}}),
            "dollar_linked":             payload.get("dollar_linked", {}),
            "hard_dollar":               payload.get("hard_dollar", {}),
            "hard_dollar_curve_history": payload.get("hard_dollar_curve_history", {"dates": [], "curves": {}}),
            "hero_metrics":              payload.get("hero_metrics", []),
        }
        js_path = os.path.join(DATA_DIR, "mercados.js")
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        js_content = (
            "// Mercados data - extraido de dashboard_payload - " + ts + "\n"
            "window.MERCADOS_DATA = "
            + json.dumps(subset, ensure_ascii=False, default=str)
            + ";\n"
        )
        with open(js_path, "w", encoding="utf-8") as f:
            f.write(js_content)
        sz = os.path.getsize(js_path)
        status.ok("Mercados", f"{sz:,} bytes - {ts}")
    except Exception as e:
        status.warn("Mercados", f"no se pudo actualizar desde la API: {e}")

    # ---- Resumen ----
    print()
    print("=" * 64)
    ok   = sum(1 for r in status.results if r[0] == "OK")
    warn = sum(1 for r in status.results if r[0] == "WARN")
    fail = sum(1 for r in status.results if r[0] == "FAIL")
    print(f"Resumen: {ok} OK - {warn} warnings - {fail} errores")
    print()
    if fail == 0:
        print("LISTO: refrescar el dashboard en el navegador.")
        return 0
    else:
        print("Hubo errores. Revisar las rutas en CONFIGURACION del script.")
        return 1

if __name__ == "__main__":
    rc = main()
    sys.exit(rc)
